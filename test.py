from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = 5.1
ws["B15"] = "here"
wb.save('data.xlsx')