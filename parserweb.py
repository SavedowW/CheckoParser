from bs4 import BeautifulSoup
import requests
import dearpygui.dearpygui as dpg
from openpyxl import Workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import re
from datetime import date

current_date = (int(date.today().day), int(date.today().month), int(date.today().year))
min_date = [current_date[0], current_date[1], current_date[2]]
max_date = [current_date[0], current_date[1], current_date[2]]
filter_by_date = False

webprefix = "https://checko.ru"
regions = ("RU", "BY")
region_links = ("https://checko.ru/company/select?code=all", "https://checko.ru/by/company/select?code=all")
target_file = "data.xlsx"
selected_region = 0
list_categories_links = []
list_categories_subcats_links = []
active_only = False
attempts = 10

country_regions = []
list_country_regions = []
selected_country_region = 0

region_cities = []
list_region_cities = []
selected_region_city = 0

month_ids = {
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12
}

def parse_date(s):
    matched = re.match(r"(\d*) (\w+) (\d{4})", s)
    if not matched or len(matched.groups()) < 3:
           return False
    month_s = matched.group(2).lower()
    if month_s in month_ids:
        return (int(matched.group(1)), month_ids[month_s], int(matched.group(3)))
    else:
        return False
    
def cmp_dates(dt1, dt2):
    if dt1[2] > dt2[2]:
        return 1
    elif dt1[2] < dt2[2]:
        return -1
    else:
        if dt1[1] > dt2[1]:
            return 1
        elif dt1[1] < dt2[1]:
            return -1
        else:
            if dt1[0] > dt2[0]:
                return 1
            elif dt1[0] < dt2[0]:
                return -1
            else:
                return 0

def getByURL(url):
    session = requests.Session()
    retry = Retry(connect=5, backoff_factor=1.0)
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    return session.get(url)

def getHolderPlaceholder(link):
    return {
        "link": link,
        "name": "-",
        "ogrn": "-",
        "inn": "-",
        "UNP": "-",
        "register_date": "-",
        "register_date_parsed": "-",
        "activity_type": "-",
        "address": "-",
        "org_type": "-",
        "capital": "-",
        "special_tax_mode": "-",
        "holder": "-",
        "avg_worker_count": "-",
        "phones": "-",
        "emails": "-",
        "current_gov": "-",
        "registrator": "-",
        "head_position": "-",
        "head_name": "-"
    }

def get_activity_categories(url):
    response = getByURL(url)
    bs = BeautifulSoup(response.text, "lxml")
    atags = bs.find_all("a", {"class": "link"})
    links = []
    for i in atags:
        if (i.get_text() == "каталогом"):
            continue
        links.append([i.get_text(), i.get('href'), False])

    return links

def save_ru_data(filename, companies):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Ссылка"
    ws["B1"] = "Название"
    ws["C1"] = "ОГРН"
    ws["D1"] = "ИНН"
    ws["E1"] = "Дата регистрации"
    ws["F1"] = "Вид деятельности"
    ws["G1"] = "Юридический адрес"
    ws["H1"] = "Организационно-правовая форма"
    ws["I1"] = "Уставный капитал"
    ws["J1"] = "Специальный налоговый режим"
    ws["K1"] = "Держатель реестра акционеров"
    ws["L1"] = "Среднесписочная численность работников"
    ws["M1"] = "Почта"
    ws["N1"] = "Телефон"
    ws["O1"] = "Должность представителя"
    ws["P1"] = "ФИО представителя"

    i = 2
    for comp in companies:
        itxt = str(i)
        ws["A"+itxt] = comp["link"]
        ws["B"+itxt] = comp["name"]
        ws["C"+itxt] = comp["ogrn"]
        ws["D"+itxt] = comp["inn"]
        ws["E"+itxt] = comp["register_date"]
        ws["F"+itxt] = comp["activity_type"]
        ws["G"+itxt] = comp["address"]
        ws["H"+itxt] = comp["org_type"]
        ws["I"+itxt] = comp["capital"]
        ws["J"+itxt] = comp["special_tax_mode"]
        ws["K"+itxt] = comp["holder"]
        ws["L"+itxt] = comp["avg_worker_count"]
        ws["M"+itxt] = comp["emails"]
        ws["N"+itxt] = comp["phones"]
        ws["O"+itxt] = comp["head_position"]
        ws["P"+itxt] = comp["head_name"]

        i += 1

    wb.save(target_file)

def save_by_data(filename, companies):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Ссылка"
    ws["B1"] = "Название"
    ws["C1"] = "УНП"
    ws["D1"] = "Дата регистрации"
    ws["E1"] = "Основной вид деятельности"
    ws["F1"] = "Юридический адрес"
    ws["G1"] = "Текущий орган учета"
    ws["H1"] = "Орган, принявший решение о регистрации"
    ws["I1"] = "Почта"
    ws["J1"] = "Телефон"

    i = 2
    for comp in companies:
        itxt = str(i)
        ws["A"+itxt] = comp["link"]
        ws["B"+itxt] = comp["name"]
        ws["C"+itxt] = comp["UNP"]
        ws["D"+itxt] = comp["register_date"]
        ws["E"+itxt] = comp["activity_type"]
        ws["F"+itxt] = comp["address"]
        ws["G"+itxt] = comp["current_gov"]
        ws["H"+itxt] = comp["registrator"]
        ws["I"+itxt] = comp["emails"]
        ws["J"+itxt] = comp["phones"]

        i += 1

    wb.save(target_file)


def add_output_message(str):
    dpg.set_value("outputMessage", dpg.get_value("outputMessage") + "\n" + str)

# parsing ru company page
def get_ru_company_data(url):
    print("CALLED RU PARSE")
    print(url)
    response = getByURL(url)
    print("Got request response: " + str(response))
    if (response.status_code != 200):
        return False
    holder = getHolderPlaceholder(url)
    print("Got placeholder")
    bs = BeautifulSoup(response.text, "lxml")
    print("Parsed")
    basicsec = bs.find_all("section", {"id": "basic"})
    if (len(basicsec) == 0):
        print("No company data")
        add_output_message("Данные по компании отсутствуют")
        return holder
    name = basicsec[0].find("p", {"class": "mb-4"})
    if (name):
        holder["name"] = name.get_text()

    datas = basicsec[0].find_all("div", {"class": "basic-data"})
    ogrnelem = basicsec[0].find("strong", {"id": "copy-ogrn"})
    if (ogrnelem):
        holder["ogrn"] = ogrnelem.get_text()
    innelem = basicsec[0].find("strong", {"id": "copy-inn"})
    if (innelem):
        holder["inn"] = innelem.get_text()
    for data in datas:
        innerdivs = data.find_all("div")
        if (len(innerdivs) >= 2):
            ttl = innerdivs[0].get_text()
            print("ttl: " + str(ttl))
            if (ttl == "Дата регистрации"):
                holder["register_date"] = innerdivs[1].get_text()
                holder["register_date_parsed"] = parse_date(holder["register_date"])
            elif (ttl == "Вид деятельности"):
                holder["activity_type"] = innerdivs[1].get_text()
            elif (ttl == "Юридический адрес"):
                holder["address"] = innerdivs[1].get_text()
            elif (ttl == "Организационно-правовая форма"):
                holder["org_type"] = innerdivs[1].get_text()
            elif (ttl == "Уставный капитал"):
                holder["capital"] = innerdivs[1].get_text()
            elif (ttl == "Специальный налоговый режим"):
                spntodelete = innerdivs[1].find_all("span")
                for i in spntodelete:
                    i.decompose()
                holder["special_tax_mode"] = innerdivs[1].get_text()
            elif (ttl == "Держатель реестра акционеров"):
                holder["holder"] = innerdivs[1].get_text()
            elif (ttl == "Среднесписочная численность работников"):
                spntodelete = innerdivs[1].find_all("span")
                for i in spntodelete:
                    i.decompose()
                holder["avg_worker_count"] = innerdivs[1].get_text()
        picscnt1 = len(data.find_all("picture"))
        picscnt2 = len(data.find_all("a", {"data-type": "image"}))
        if (picscnt1 > 0 or picscnt2 > 0):
            titlediv = data.find("div", {"class": "uk-text-bold"})
            namea = data.find("a", {"class": "link"})
            if (titlediv and namea):
                holder["head_position"] = titlediv.get_text()
                holder["head_name"] = namea.get_text()

    contactssec = bs.find_all("section", {"id": "contacts"})
    if (len(contactssec) > 0):
        contactsdivs = contactssec[0].find_all("div", {"class": "uk-grid-divider"})[0].find_all("div", {"class": "uk-width-1"})
        phones = contactsdivs[0].find_all("a", {"class": "black-link"})
        phones_str = ""
        for i in phones:
            if (len(phones_str) > 0):
                phones_str += ", "
            phones_str += i.get_text()
        if (len(phones_str) > 0):
            holder["phones"] = phones_str
        else:
            holder["phones"] = "-"
        emails = contactsdivs[1].find_all("a", {"class": "link"})
        emails_str = ""
        for i in emails:
            email_s = i.get_text()
            if email_s.__contains__("@"):
                if (len(emails_str) > 0):
                    emails_str += ", "
                emails_str += email_s
        if (len(emails_str) > 0):
            holder["emails"] = emails_str
        else:
            holder["emails"] = "-"
    print("Ended parsing")
    return holder

# parsing by company page
def get_by_company_data(url):
    print("CALLED BY PARSE")
    print(url)
    response = getByURL(url)
    print("Got request response: " + str(response))
    if (response.status_code != 200):
        return False
    holder = getHolderPlaceholder(url)
    print("Got placeholder")
    bs = BeautifulSoup(response.text, "lxml")
    print("Parsed")
    basicsec = bs.find_all("section", {"id": "basic"})
    if (len(basicsec) == 0):
        print("No company data")
        add_output_message("Данные по компании отсутствуют")
        return holder
    name = basicsec[0].find("p", {"class": "mb-4"})
    if (name):
        holder["name"] = name.get_text()

    col_datas = basicsec[0].find_all("div", {"class": "basic-data"})
    unpelem = col_datas[1].find("strong", {"id": "copy-id"})
    if (unpelem):
        holder["UNP"] = unpelem.get_text()
    for data in col_datas:
        innerdivs = data.find_all("div")
        if (len(innerdivs) >= 1):
            ttl = innerdivs[0].get_text()
            print("ttl: " + str(ttl))
            print(ttl)
            if (ttl == "Дата регистрации"):
                holder["register_date"] = innerdivs[1].get_text()
                holder["register_date_parsed"] = parse_date(holder["register_date"])
            if (ttl == "Основной вид деятельности"):
                holder["activity_type"] = innerdivs[1].get_text()
            if (ttl == "Юридический адрес"):
                holder["address"] = data.find_all("strong")[0].get_text()
                print("Found address: " + holder["address"])
            if (ttl == "Текущий орган учёта"):
                holder["current_gov"] = innerdivs[1].get_text()

    regsec = bs.find_all("section", {"id": "registration"})
    col1_content = regsec[0].find_all("div", {"class": "uk-width-1"})[0]
    col1_divs = col1_content.find_all("div")
    holder["registrator"] = col1_divs[3].get_text()

    consec = bs.find_all("section", {"id": "contacts"})
    if (len(consec) > 0):
        grid = consec[0].find_all("div", {"class": "uk-grid-divider"})[0].find_all("div", {"class": "uk-width-1"})
        emails = grid[1].find_all("a", {"class": "link"})
        emails_str = ""
        for i in emails:
            email_s = i.get_text()
            print(email_s)
            if email_s.__contains__("@"):
                if (len(emails_str) > 0):
                    emails_str += ", "
                emails_str += email_s
        if (len(emails_str) > 0):
            holder["emails"] = emails_str
        else:
            holder["emails"] = "-"

        phones = grid[0].find_all("a", {"class": "black-link"})
        phones_str = ""
        for i in phones:
            phone_s = i.get_text()
            if (len(phones_str) > 0):
                phones_str += ", "
            phones_str += phone_s
        if (len(phones_str) > 0):
            holder["phones"] = phones_str
        else:
            holder["phones"] = "-"
    print("Ended parsing")
    return holder

def get_subcat_links(url):
    response = getByURL(url)
    bs = BeautifulSoup(response.text, "lxml")
    subcatsdiv = bs.find_all("div", {"class": "sub-okveds"})
    if (len(subcatsdiv) == 0):
        subcatsdiv = bs.find_all("div", {"class": "sub-okeds"})
    if (len(subcatsdiv) == 0):
        return []
    atags = subcatsdiv[0].find_all("a", {"class": "link"})
    lst = []
    for atag in atags:
        lst.append([atag["href"], atag.get_text(), False])
        lst = lst + get_subcat_links(webprefix + atag["href"])
    return lst

# parsing single page
def parse_single_companies_page(url, isRu):
    print("Parsing url: " + url)
    add_output_message("Обработка страницы: " + url)
    response = getByURL(url)
    bs = BeautifulSoup(response.text, "lxml")
    atags = bs.find_all("td", {"class": ""})
    lst = []
    for i in atags:
        company = i.find("a", {"class": "link"})
        if (company != None):
            add_output_message("Обработка компании " + company.get_text())
            cmpurl = webprefix + company["href"]
            attempts_count = attempts
            compres = False
            while (attempts_count > 0 and compres == False):
                compres = get_ru_company_data(cmpurl) if isRu else get_by_company_data(cmpurl)
                attempts_count -= 1
            if compres == False:
                add_output_message("Не получилось получить данные о компании")
            else:
                if not compres["register_date_parsed"]:
                    add_output_message("Не получилось обработать дату регистрации")
                else:
                    date = compres["register_date_parsed"]
                    if cmp_dates(date, min_date) != -1 and cmp_dates(date, max_date) != 1:
                        lst.append(compres)
                    else:
                        add_output_message("Не выполняется условие по дате")
    
    return lst

# parsing all pages from baseurl
def parse_companies_pages(lst, baseurl, isRu):
    if (active_only):
        baseurl += "&active=true"
    if (isRu):
        regid = country_regions[selected_country_region][0]
        if (regid != -1):
            cityid = region_cities[selected_region_city][0]
            if (cityid != -1):
                baseurl += "&city=" + str(cityid)
            else:
                baseurl += "&region=" + str(regid)
    res = parse_single_companies_page(baseurl, isRu)
    lst = lst + res
    i = 1
    while (len(res) > 0):
        i += 1
        add_output_message("Обработка страницы " + str(i))
        url = baseurl + "&page=" + str(i)
        res = parse_single_companies_page(url, isRu)
        lst = lst + res
    return lst

def get_regions_list():
    url = webprefix + list_categories_links[0][1]
    print(url)
    response = getByURL(url)
    bs = BeautifulSoup(response.text, "lxml")
    divs = bs.find_all("div", {"class": "data-select-dropdown"})
    regionrawlst = divs[0].find_all("li")
    lst = []
    for el in regionrawlst:
        elstr = str(el)
        name = el.get_text()
        num = 0
        if ("region_clear" in str(el)):
            num = -1
        else:
            res = re.findall(r'(?<=region_select\(\')\d+(?=\'\))', elstr)
            num = int(res[0])
        lst.append([num, name])
            
    return lst

def get_cities_list():
    regid = country_regions[selected_country_region][0]
    url = webprefix + list_categories_links[0][1]
    if (regid != -1):
        url += "&region=" + str(regid)
    print(url)
    response = getByURL(url)
    bs = BeautifulSoup(response.text, "lxml")
    divs = bs.find_all("div", {"class": "data-select-dropdown"})
    regionrawlst = divs[1].find_all("li")
    lst = []
    for el in regionrawlst:
        elstr = str(el)
        name = el.get_text()
        num = 0
        print(el)
        if ("city_clear" in str(el)):
            num = -1
        else:
            res = re.findall(r'(?<=city_select\()\d+(?=\))', elstr)
            num = int(res[0])
        lst.append([num, name])
            
    return lst

def select_region(regname):
    global selected_region, list_categories_links, country_regions, list_country_regions, selected_country_region, region_cities, list_region_cities, selected_region_city
    selected_region = regions.index(regname)
    add_output_message("Получаем список категорий...")
    list_categories_links = get_activity_categories(region_links[selected_region])
    if (regname == 'RU'):
        add_output_message("Получаем список регионов...")
        country_regions = get_regions_list()
        add_output_message("Данные загружены.")
        list_country_regions = [k for i, k in country_regions]
        selected_country_region = 0
        dpg.configure_item("regionscombo", items=list_country_regions, default_value=list_country_regions[selected_country_region])
        region_cities = []
        list_region_cities = []
        selected_region_city = 0
        dpg.configure_item("citiescombo", items=list_region_cities, default_value="")
    else:
        add_output_message("Данные загружены.")
        country_regions = []
        list_country_regions = []
        selected_country_region = 0
        dpg.configure_item("regionscombo", items=list_country_regions, default_value="")
        region_cities = []
        list_region_cities = []
        selected_region_city = 0
        dpg.configure_item("citiescombo", items=list_region_cities, default_value="")

def select_country_region(sender, app_data):
    global selected_country_region, region_cities, list_region_cities, selected_region_city
    print(sender)
    print(app_data)
    selected_country_region = list_country_regions.index(app_data)
    print(selected_country_region)
    add_output_message("Получаем список городов...")
    region_cities = get_cities_list()
    add_output_message("Данные загружены.")
    list_region_cities = [k for i, k in region_cities]
    selected_region_city = 0
    dpg.configure_item("citiescombo", items=list_region_cities, default_value=list_region_cities[selected_region_city])

def select_region_city(sender, app_data):
    global selected_region_city
    selected_region_city = list_region_cities.index(app_data)
    print(selected_region_city)

def callback_selectable(sender, app_data, user_data):
    global list_categories_links
    list_categories_links[user_data][2] = app_data

def callback_subcat_selectable(sender, app_data, user_data):
    global list_categories_subcats_links
    list_categories_subcats_links[user_data][2] = app_data

def callback_select_country(sender, app_data):
    global list_categories_links, selected_region
    select_region(app_data)

    dpg.delete_item("cattbl", children_only=True, slot=1)
    i = 0
    s = ""
    for el in list_categories_links:
        s = "blbl" + str(i)
        dpg.add_table_row(label="here2", parent="cattbl", tag="blbl" + str(i))
        dpg.add_selectable(label=el[0], parent=s, tag = s + "el", user_data=(i), callback=callback_selectable)
        i += 1

def callback_select_file(sender, app_data):
    global target_file
    target_file = app_data["file_path_name"]
    dpg.set_value("selectedFile", "Текущий файл: " + target_file)

def callback_active_only(sender, app_data, user_data):
    global active_only
    active_only = app_data

def callback_date_filter_checkbox(sender, app_data, user_data):
    global filter_by_date
    if app_data:
        dpg.show_item("day_min_filter")
        dpg.show_item("month_min_filter")
        dpg.show_item("year_min_filter")
        dpg.show_item("day_max_filter")
        dpg.show_item("month_max_filter")
        dpg.show_item("year_max_filter")
    else:
        dpg.hide_item("day_min_filter")
        dpg.hide_item("month_min_filter")
        dpg.hide_item("year_min_filter")
        dpg.hide_item("day_max_filter")
        dpg.hide_item("month_max_filter")
        dpg.hide_item("year_max_filter")

    filter_by_date = app_data

def callback_date_filter(sender, app_data, user_data):
    global min_date, max_date
    if sender == "day_min_filter":
        min_date[0] = int(app_data)
    elif sender == "day_max_filter":
        max_date[0] = int(app_data)
    elif sender == "month_min_filter":
        min_date[1] = int(app_data)
    elif sender == "month_max_filter":
        max_date[1] = int(app_data)
    elif sender == "year_min_filter":
        min_date[2] = int(app_data)
    elif sender == "year_max_filter":
        max_date[2] = int(app_data)

def callback_parse(sender, app_data):
    if (len(list_categories_links) == 0):
        return
    
    try:
        add_output_message("Начат парсинг со следующими настройками:")
        add_output_message("Регион: " + regions[selected_region])
        if filter_by_date:
            add_output_message("Включен фильтр по датам")
            add_output_message("От: " + str(min_date))
            add_output_message("До: " + str(max_date))

        add_output_message("Выбранные виды деятельности: ")
        selected = []
        for i in range(len(list_categories_links)):
            if list_categories_links[i][2]:
                selected.append(i)
                add_output_message(list_categories_links[i][0])

        add_output_message("Выбранные подкатегории: ")
        selected_subcats = []
        for i in range(len(list_categories_subcats_links)):
            if list_categories_subcats_links[i][2]:
                selected_subcats.append(i)
                add_output_message(list_categories_subcats_links[i][1])

        print(selected)
        lst = []
        for i in selected:
            add_output_message("Текущая категория: " + list_categories_links[i][0])
            lst = parse_companies_pages(lst, webprefix + list_categories_links[i][1], selected_region == 0)

        for i in selected_subcats:
            add_output_message("Текущая подкатегория: " + list_categories_subcats_links[i][1])
            lst = parse_companies_pages(lst, webprefix + list_categories_subcats_links[i][0], selected_region == 0)

        add_output_message("Парсинг окончен, сохраняем в файл")
        if (selected_region == 0):
            save_ru_data(target_file, lst)
        else:
            save_by_data(target_file, lst)
    except Exception as ex:
        add_output_message("Возникла неопознанная ошибка")
        add_output_message(str(ex))

def callback_update_subcategory_list(sender, app_data):
    global list_categories_subcats_links
    add_output_message("Обновляем список подкатегорий...")
    dpg.delete_item("subcattbl", children_only=True, slot=1)
    list_categories_subcats_links = []
    for i in range(len(list_categories_links)):
        if list_categories_links[i][2]:
            list_categories_subcats_links = list_categories_subcats_links + get_subcat_links(webprefix + list_categories_links[i][1])
    i = 0
    for el in list_categories_subcats_links:
        dpg.add_table_row(label="here2", parent="subcattbl", tag="subct" + str(i))
        dpg.add_selectable(label=el[1], parent="subct" + str(i), tag="subctel" + str(i), user_data=(i), callback=callback_subcat_selectable)
        i += 1
    add_output_message("Список подкатегорий обновлен")

def callback_select_all_categories(sender, app_data):
    global list_categories_links
    for i in range(0, len(list_categories_links)):
        dpg.set_value("blbl" + str(i) + "el", True)
        list_categories_links[i][2] = True

def callback_select_all_subcategories(sender, app_data):
    global list_categories_subcats_links
    for i in range(0, len(list_categories_subcats_links)):
        dpg.set_value("subctel" + str(i), True)
        list_categories_subcats_links[i][2] = True

dpg.create_context()

# Font from https://fonts-online.ru/fonts/noto-mono
with dpg.font_registry():
    with dpg.font("notomono-regular.ttf", 13, default_font=True, tag="Default font") as f:
        dpg.add_font_range_hint(dpg.mvFontRangeHint_Cyrillic)
dpg.bind_font("Default font")

dpg.create_viewport(title='Custom Title', width=1660, height=720)
with dpg.file_dialog(directory_selector=False, show=False, callback=callback_select_file, id="file_dialog_id", width=500, height=400):
        dpg.add_file_extension("Source files (*.xlsx *.xls){.xlsx,.xls}", color=(0, 255, 255, 255))

with dpg.window(label="Select categories", tag="categories_selector", width=618, height=677, pos=(404, 2)):
    with dpg.table(header_row=True, tag="cattbl"):
        dpg.add_table_column(label="Категории")
    dpg.add_button(label="Выбрать все", callback=callback_select_all_categories)

with dpg.window(label="Select subcategories", tag="subcategories_selector", width=618, height=300, pos=(1024, 2)):
    dpg.add_button(label="Обновить подкатегории", callback=callback_update_subcategory_list)
    with dpg.table(header_row=True, tag="subcattbl"):
        dpg.add_table_column(label="Подкатегории")
    dpg.add_button(label="Выбрать все", callback=callback_select_all_subcategories)

with dpg.window(label="Select region (RU only)", tag="region_selector", width=618, height=375, pos=(1024, 304)):
    dpg.add_combo((), label="Регионы", tag="regionscombo", callback=select_country_region)
    dpg.add_combo((), label="Города", tag="citiescombo", callback=select_region_city)

with dpg.window(label="Checko parser", width=400, height=250, pos=(2, 2)):
    dpg.add_text("Выберите страну и вид деятельности")
    dpg.add_combo(("RU", "BY"), label="Страна", callback=callback_select_country)
    dpg.add_button(label="Выберите файл", callback=lambda: dpg.show_item("file_dialog_id"))
    dpg.add_text(default_value="Текущий файл: " + target_file, tag="selectedFile")
    dpg.add_checkbox(label="Только действующие организации", callback=callback_active_only)
    dpg.add_checkbox(label="Включить фильтр по дате регистрации", callback=callback_date_filter_checkbox)
    dpg.add_combo([i for i in range(1, 32)], label="День", default_value=current_date[0], show=False, width=50, tag="day_min_filter", callback=callback_date_filter)
    dpg.add_same_line()
    dpg.add_combo([i for i in range(1, 13)], label="Месяц", default_value=current_date[1], show=False, width=50, tag="month_min_filter", callback=callback_date_filter)
    dpg.add_same_line()
    dpg.add_combo([i for i in range(1970, current_date[2] + 1)], label="Год (минимум)", default_value=current_date[2], show=False, width=78, tag="year_min_filter", callback=callback_date_filter)
    dpg.add_combo([i for i in range(1, 32)], label="День", default_value=current_date[0], show=False, width=50, tag="day_max_filter", callback=callback_date_filter)
    dpg.add_same_line()
    dpg.add_combo([i for i in range(1, 13)], label="Месяц", default_value=current_date[1], show=False, width=50, tag="month_max_filter", callback=callback_date_filter)
    dpg.add_same_line()
    dpg.add_combo([i for i in range(1970, current_date[2] + 1)], label="Год (максимум)", default_value=current_date[2], show=False, width=78, tag="year_max_filter", callback=callback_date_filter)
    dpg.add_button(label="СТАРТ", callback=callback_parse)

with dpg.window(label="Log", width=400, height=424, pos=(2, 255)):
    dpg.add_text(default_value="",tag="outputMessage")

dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()
dpg.destroy_context()